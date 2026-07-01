import json
from collections import Counter, OrderedDict
from datetime import date, datetime, time

from django.conf import settings
from django.http import HttpResponse
from django.utils.encoding import force_str
from django.utils.functional import Promise
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from common.drf.renders.excel import ExcelFileRenderer
from common.drf.renders.mixins import LogMixin


class BaseListReportExporter:
    format = 'xlsx'
    data_sheet_name = _('Data')
    summary_sheet_name = _('Summary')
    report_title = _('Report')
    report_basename = ''
    max_rows = settings.MAX_LIMIT_PER_PAGE
    min_column_width = 12
    max_column_width = 48

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    section_fill = PatternFill('solid', fgColor='EAF4EA')
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    filter_label_map = {
        'date_from': _('Date from'),
        'date_to': _('Date to'),
    }

    def __init__(self, *, view, request, queryset):
        self.view = view
        self.request = request
        self.queryset = queryset
        self.records = self._materialize_records(queryset)
        self.serializer = self.view.get_serializer(self.records, many=True)
        self.export_serializer = getattr(self.serializer, 'child', self.serializer)
        self.render_helper = self._get_render_helper()
        self.serialized_rows = self.render_helper.process_data(self.serializer.data)

    def _materialize_records(self, queryset):
        if isinstance(queryset, list):
            return queryset[:self.max_rows]
        return list(queryset[:self.max_rows])

    def _get_render_helper(self):
        helper = ExcelFileRenderer()
        helper.serializer = self.export_serializer
        helper.template = 'export'
        return helper

    def _sanitize_value(self, value):
        if value is None:
            return '-'
        if isinstance(value, Promise):
            value = force_str(value)
        elif not isinstance(value, (str, int, float, bool, datetime, date, time)):
            value = force_str(value)
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub('', value)
        return value

    def _set_text_cell(self, cell, value):
        if isinstance(value, str):
            cell.value = self._sanitize_value(value)
            cell.data_type = 's'
        else:
            cell.value = value

    def _write_row(self, worksheet, row_index, values, *, bold=False, fill=None):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            self._set_text_cell(cell, self._sanitize_value(value))
            cell.alignment = self.wrap_alignment
            if bold:
                cell.font = self.header_font
            if fill is not None:
                cell.fill = fill
        return row_index + 1

    def _write_section_title(self, worksheet, row_index, title, width):
        width = max(width, 2)
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=width)
        cell = worksheet.cell(row=row_index, column=1)
        self._set_text_cell(cell, str(title))
        cell.font = self.section_font
        cell.fill = self.section_fill
        cell.alignment = self.wrap_alignment
        return row_index + 1

    def _auto_fit(self, worksheet):
        for cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(cells[0].column)
            for cell in cells:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            width = max(self.min_column_width, min(max_length + 2, self.max_column_width))
            worksheet.column_dimensions[column_letter].width = width

    def _format_query_params(self):
        params = LogMixin._clean_params(self.request.query_params)
        if not params:
            return _('All data')
        items = []
        for key, value in params.items():
            if isinstance(value, list):
                value = ', '.join([str(item) for item in value])
            elif isinstance(value, (dict, OrderedDict)):
                value = json.dumps(value, ensure_ascii=False)
            label = self.get_filter_label(key)
            items.append(f'{label}={value}')
        return '; '.join(items)

    def get_filter_label(self, key):
        label = self.filter_label_map.get(key)
        if label is not None:
            return force_str(label)
        fields = getattr(self.export_serializer, 'fields', {})
        field = fields.get(key)
        if field is not None and getattr(field, 'label', None):
            return force_str(field.label)
        return key

    def get_report_basename(self):
        if self.report_basename:
            return self.report_basename
        model = getattr(self.view, 'model', None)
        if model is not None:
            return f'{model.__name__.lower()}_report'
        return 'report'

    def get_filename(self):
        suffix = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S')
        return f'{self.get_report_basename()}_{suffix}.{self.format}'

    def get_response(self):
        workbook = self.build_workbook()
        content = save_virtual_workbook(workbook)
        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename()}"'
        return response

    def record_logs(self):
        LogMixin().record_logs(self.request, self.view, self.serialized_rows)

    def get_data_sheet_headers(self):
        fields = self.render_helper.get_rendered_fields()
        headers = [str(title) for title in self.render_helper.get_column_titles(fields)]
        return fields, headers

    def get_data_sheet_rows(self):
        fields, headers = self.get_data_sheet_headers()
        rows = []
        for row in self.serialized_rows:
            current = []
            for field in fields:
                field._row = row
                current.append(self.render_helper.render_value(field, row.get(field.field_name)))
            rows.append(current)
        return headers, rows

    def get_summary_metadata(self):
        return OrderedDict([
            (_('Report'), self.report_title),
            (_('Generated at'), timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')),
            (_('Exported rows'), len(self.records)),
            (_('Filters'), self._format_query_params()),
        ])

    def get_summary_sections(self):
        return []

    def build_workbook(self):
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = str(self.summary_sheet_name)
        self.build_summary_sheet(summary_sheet)

        data_sheet = workbook.create_sheet(title=str(self.data_sheet_name))
        self.build_data_sheet(data_sheet)
        return workbook

    def build_data_sheet(self, worksheet):
        headers, rows = self.get_data_sheet_rows()
        next_row = self._write_row(worksheet, 1, headers, bold=True, fill=self.header_fill)
        for row in rows:
            next_row = self._write_row(worksheet, next_row, row)
        worksheet.freeze_panes = 'A2'
        self._auto_fit(worksheet)

    def build_summary_sheet(self, worksheet):
        row_index = 1
        metadata = list(self.get_summary_metadata().items())
        row_index = self._write_section_title(worksheet, row_index, _('Report Summary'), 2)
        for key, value in metadata:
            row_index = self._write_row(worksheet, row_index, [key, value])
        row_index += 1

        for section in self.get_summary_sections():
            section_type = section.get('type')
            title = section.get('title', '')
            if section_type == 'key_value':
                rows = list(section.get('rows', []))
                row_index = self._write_section_title(worksheet, row_index, title, 2)
                for key, value in rows:
                    row_index = self._write_row(worksheet, row_index, [key, value])
            elif section_type == 'table':
                headers = list(section.get('headers', []))
                rows = list(section.get('rows', []))
                row_index = self._write_section_title(worksheet, row_index, title, len(headers))
                if headers:
                    row_index = self._write_row(worksheet, row_index, headers, bold=True, fill=self.header_fill)
                for values in rows:
                    row_index = self._write_row(worksheet, row_index, values)
            row_index += 1

        self._auto_fit(worksheet)

    @staticmethod
    def build_key_value_section(title, rows):
        return {
            'type': 'key_value',
            'title': title,
            'rows': list(rows),
        }

    @staticmethod
    def build_table_section(title, headers, rows):
        return {
            'type': 'table',
            'title': title,
            'headers': list(headers),
            'rows': list(rows),
        }

    @staticmethod
    def build_counter_rows(counter, *, top_n=None):
        items = list(counter.items())
        items.sort(key=lambda item: (-item[1], str(item[0])))
        if top_n is not None:
            items = items[:top_n]
        return [[key, value] for key, value in items]

    @staticmethod
    def format_percent(numerator, denominator):
        if not denominator:
            return '0.00%'
        return '{:.2f}%'.format((numerator / denominator) * 100)

    @staticmethod
    def local_date_string(value):
        if isinstance(value, datetime):
            value = timezone.localtime(value)
        else:
            value = timezone.localtime(value)
        return value.strftime('%Y-%m-%d')

    @staticmethod
    def build_top_counter(values):
        counter = Counter()
        for value in values:
            if value in ('', None):
                continue
            counter[str(value)] += 1
        return counter
